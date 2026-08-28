import csv
import re
import sqlite3
from pathlib import Path

from rag.paths import DB_PATH, STORE_DIR

ROW_LIMIT_SAMPLE = 3


def safe_table_name(*parts):
    merged = "_".join(str(part) for part in parts if part)
    name = re.sub(r"[^a-zA-Z0-9]+", "_", merged).strip("_").lower()
    if not name:
        name = "table"
    if name[0].isdigit():
        name = f"t_{name}"
    return name


def connect_db(readonly=False):
    if readonly:
        if not DB_PATH.exists():
            raise FileNotFoundError(
                "No structured data index yet. Run: python -m rag.ingest"
            )
        uri = f"file:{DB_PATH.as_posix()}?mode=ro"
        conn = sqlite3.connect(uri, uri=True)
    else:
        STORE_DIR.mkdir(parents=True, exist_ok=True)
        conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def reset_db():
    if DB_PATH.exists():
        DB_PATH.unlink()
    return connect_db(readonly=False)


def import_csv(conn, path: Path):
    table = safe_table_name(path.stem)
    with path.open(encoding="utf-8", newline="", errors="replace") as handle:
        reader = csv.DictReader(handle)
        columns = list(reader.fieldnames or [])
        rows = list(reader)
    _create_and_fill(conn, table, columns, rows, path.name)
    return table


def import_excel(conn, path: Path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    tables = []
    for sheet in workbook.worksheets:
        rows_iter = sheet.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        columns = [
            str(name).strip() if name is not None else f"col_{index}"
            for index, name in enumerate(header, start=1)
        ]
        records = []
        for row in rows_iter:
            record = {}
            for index, column in enumerate(columns):
                value = row[index] if index < len(row) else ""
                record[column] = "" if value is None else str(value)
            if any(str(value).strip() for value in record.values()):
                records.append(record)
        table = safe_table_name(path.stem, sheet.title)
        _create_and_fill(conn, table, columns, records, f"{path.name}:{sheet.title}")
        tables.append(table)
    return tables


def import_sqlite(conn, path: Path):
    prefix = safe_table_name(path.stem)
    source = sqlite3.connect(path)
    source.row_factory = sqlite3.Row
    names = source.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
    ).fetchall()
    imported = []
    for (name,) in names:
        dest = safe_table_name(prefix, name)
        quoted = name.replace('"', '""')
        rows = source.execute(f'SELECT * FROM "{quoted}"').fetchall()
        if rows:
            columns = list(rows[0].keys())
            records = [dict(row) for row in rows]
        else:
            columns = [
                info[1]
                for info in source.execute(f'PRAGMA table_info("{quoted}")').fetchall()
            ]
            records = []
        _create_and_fill(conn, dest, columns, records, f"{path.name}:{name}")
        imported.append(dest)
    source.close()
    return imported


def list_table_cards(conn):
    names = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' "
        "AND name NOT LIKE 'sqlite_%' AND name != '_sources' ORDER BY name"
    ).fetchall()
    cards = []
    for (name,) in names:
        columns = [
            row[1] for row in conn.execute(f'PRAGMA table_info("{name}")').fetchall()
        ]
        sample = conn.execute(
            f'SELECT * FROM "{name}" LIMIT {ROW_LIMIT_SAMPLE}'
        ).fetchall()
        sample_lines = [
            ", ".join(f"{key}={row[key]}" for key in row.keys()) for row in sample
        ]
        origin_row = conn.execute(
            "SELECT origin FROM _sources WHERE table_name = ?", (name,)
        ).fetchone()
        origin_label = origin_row["origin"] if origin_row else name
        text = (
            f"Table {name} (from {origin_label})\n"
            f"Columns: {', '.join(columns)}\n"
            f"Sample rows:\n"
            + ("\n".join(sample_lines) if sample_lines else "(empty)")
        )
        cards.append(
            {
                "source": "table",
                "title": origin_label,
                "locator": f"sql:{name}",
                "text": text,
                "table": name,
            }
        )
    return cards


def _create_and_fill(conn, table, columns, rows, origin):
    conn.execute(
        "CREATE TABLE IF NOT EXISTS _sources (table_name TEXT PRIMARY KEY, origin TEXT)"
    )
    clean_columns = []
    for index, column in enumerate(columns, start=1):
        name = safe_table_name(column) or f"col_{index}"
        if name in clean_columns:
            name = f"{name}_{index}"
        clean_columns.append(name)
    if not clean_columns:
        return

    defs = ", ".join(f'"{column}" TEXT' for column in clean_columns)
    conn.execute(f'DROP TABLE IF EXISTS "{table}"')
    conn.execute(f'CREATE TABLE "{table}" ({defs})')
    col_sql = ", ".join(f'"{column}"' for column in clean_columns)
    placeholders = ", ".join("?" for _ in clean_columns)
    values = []
    for row in rows:
        values.append(
            [
                "" if row.get(original) is None else str(row.get(original, ""))
                for original in columns
            ]
        )
    if values:
        conn.executemany(
            f'INSERT INTO "{table}" ({col_sql}) VALUES ({placeholders})',
            values,
        )
    conn.execute(
        "INSERT OR REPLACE INTO _sources(table_name, origin) VALUES (?, ?)",
        (table, origin),
    )
    conn.commit()
